# parses running spreadsheet and does basic analysis

# import data analysis modules
import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from datetime import time
import tkinter as tk

# variable to store spreadsheet location
loc = "/Users/ishanshah/Documents/Programming/Projects/runkeeper/running.xlsx"

# open excel spreadsheet
workbook = openpyxl.load_workbook(loc)
sheet = workbook["main"]

# calculate distance run
def getDist():
    sum = 0
    for i in range(2, len(sheet['B']) + 1):
        sum += sheet.cell(row = i, column = 2).value
    return str(round(sum, 2))

# calculate average pace
def getAvgPace():
    sum = 0
    ct = 0
    for i in range(2, len(sheet['B']) + 1):
        sum += minToSecs(calcPace(sheet.cell(row = i, column = 3).value, \
               sheet.cell(row = i, column = 2).value))
        ct += 1
    return secsToMin(int(sum / ct))

# convert minutes to seconds
def minToSecs(time):
    time = str(time)
    colon = time.index(":")
    mins = time[:colon]
    secs = time[colon + 1:]
    return int(mins) * 60 + int(secs)

# convert seconds to minutes
def secsToMin(secs):
    mins = secs / 60
    secs = secs % 60
    return str(mins) + ":" + str(secs)
    
# calculate pace given distance and duration
# uses minToSec() and secToMin() to do division with time
def calcPace(duration, distance):
    secs = minToSecs(duration)
    pace = int(secs / distance)
    return secsToMin(pace)

def main():
    execfile("app.py")
    # print "\nWelcome to runkeeper.\n"
    # print "You've run " + getDist() + " total miles."
    # print "Your average pace is " + getAvgPace() + " minutes per mile.\n"
main()