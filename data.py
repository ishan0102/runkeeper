# this module contains I/O functions for the excel spreadsheet
# and various graphing features

# import excel reading and graph making modules
import openpyxl
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
from running import *

# open excel spreadsheet 
workbook = openpyxl.load_workbook("running.xlsx")
sheet = workbook["main"]

# check if data is valid
def screenData(date, distance, duration):
    if len(date) != 8 or date[2] != '/' and date[5] != '/':
        return False

    try:
        distance = float(distance)
    except:
        return False

    if len(duration) != 5 or duration[2] != ':':
        return False

    return True

# input data into spreadsheet
def inputData(date, distance, duration, sheet):
    currentRow = findEmpty(sheet)
    cell1 = sheet.cell(row = currentRow, column = 1)
    cell2 = sheet.cell(row = currentRow, column = 2)
    cell3 = sheet.cell(row = currentRow, column = 3)
    cell4 = sheet.cell(row = currentRow, column = 4)

    cell1.value = date
    cell2.value = distance
    cell3.value = duration
    cell4.value = calcPace(distance, duration)

    cell1.alignment = Alignment(horizontal='right')
    cell2.alignment = Alignment(horizontal='right')
    cell3.alignment = Alignment(horizontal='right')
    cell4.alignment = Alignment(horizontal='right')

# find the next empty row in the sheet
def findEmpty(sheet):
    i = 1
    while (sheet.cell(row = i, column = 1).value != None):
        i += 1
    return i

# called when total distance button is pressed
def grab_dist_text():
    return "You have run a total of\n" + getDist(sheet) + " miles so far. \nKeep it up!"

# called when average pace button is pressed
def grab_avgpace_text():
    return "Your average pace is\n" + getAvgPace(sheet) + " minutes per mile. \nAwesome work!"

# called when total time button is pressed
def grab_totaltime_text():
    return "Your total time\nspent running is\n" + getTotalTime(sheet) + ". Solid!"